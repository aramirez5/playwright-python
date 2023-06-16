import openpyxl
from functions import Global_Functions
from config_test import set_up_excel

# Global variables
time_wait = 0.5
path = "Images/"
pdf = "C:/laragon/www/playwright-python/Students/Examples/Documents/sample.pdf"
excel = "C:/laragon/www/playwright-python/Students/Examples/Excel/data.xlsx"
registers = 7

# Read the excel file
file = openpyxl.load_workbook(excel)
def row_number(sheet):
    rw = file[sheet]
    return rw.max_row

def column_data(sheet, row, column):
    rw = file[sheet]
    col = rw.cell(int(row), int(column))
    return col.value

print(row_number("data"))
print(column_data("data", 3, 4))

def test_excel(set_up_excel) -> None:
    
    page = set_up_excel # ruff: noqa
    gf = Global_Functions(page)

    gf.Check_title("Datos Personales | TestingQaRvn", time_wait)

    gf.Wait(time_wait)

    for n in range (2, registers):
        name = column_data("data", n, 1)
        lastname = column_data("data", n, 2)
        email = column_data("data", n, 3)
        phone = column_data("data", n, 4)
        address = column_data("data", n, 5)

        gf.Text("//input[@id='wsf-1-field-21']", name, time_wait)
        gf.Text("//input[@id='wsf-1-field-22']", lastname, time_wait)
        gf.Text("//input[@id='wsf-1-field-23']", email, time_wait)
        gf.Text("//input[@id='wsf-1-field-24']", str(phone), time_wait)
        gf.Text("//textarea[@id='wsf-1-field-28']", address, time_wait)
        gf.Click("//button[@id='wsf-1-field-27']", time_wait)
        gf.Wait(time_wait)
        gf.Check_text(
            "//p[contains(text(),'Gracias por tu encuesta.')]",
            "Gracias",
            time_wait
        )
        page.goto("https://testingqarvn.com.es/datos-personales/")
        gf.Wait(time_wait)